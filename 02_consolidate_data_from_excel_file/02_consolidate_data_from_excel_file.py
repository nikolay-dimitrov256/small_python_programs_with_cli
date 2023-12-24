from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from calendar import month_name
import os


def find_xlsx_file():
    for filename in os.listdir('./'):
        if not os.path.isfile(filename):
            continue

        if filename.split('.')[1] == 'xlsx':
            return filename


def save_file(wb, original_name):
    original_name = original_name.split('.')[0]
    current_time = datetime.now()
    day = current_time.day
    month = month_name[current_time.month][:3].upper()
    year = str(current_time.year)[2:]
    new_name = f'{original_name}_consolidated_{day}{month}{year}.xlsx'

    wb.save(new_name)


def write_data(wb: Workbook, total_data_: list):
    with open('log.txt', 'w') as file:
        for row in total_data_:
            coma = ',' if len(row) == 6 else ''
            file.write(f"{','.join(row)}{coma}\n")

    for index, row in enumerate(total_data_):
        if index > 0:
            row[4] = int(row[4])
            row[5] = int(row[5])

        wb['Data'].append(row)


def get_merged_cells_value(wb: Workbook, sheet: str, col) -> int:
    ws = wb[sheet]
    lookup_cell = f'{col}1'

    for range_ in ws.merged_cells.ranges:
        if range_.__contains__(lookup_cell):
            col_letter = get_column_letter(range_.min_col)
            result = ws[f'{col_letter}1'].value
            return result


def process_data(wb: Workbook, sheet: str, total_data_: list):
    worksheet = wb[sheet]
    row = 3
    empty_rows = 0

    while empty_rows < 4:
        value_a = worksheet[f'A{row}'].value
        value_b = worksheet[f'B{row}'].value
        value_c = worksheet[f'C{row}'].value

        if value_a or value_b or value_c:
            empty_rows = 0
            orders_row = row + 2

            for col in 'EFGHIJKLMNOPQRST':
                order = worksheet[f'{col}{orders_row}'].value

                if order == '-' or order == 0 or order is None:
                    continue

                data = []
                region = value_a
                fbpn = value_b
                type_ = value_c
                data.extend([region, fbpn, type_])

                month = worksheet[f'{col}2'].value
                year = worksheet[f'{col}1'].value  # Does not get the value of the merged cells

                if year is None:
                    year = get_merged_cells_value(wb, sheet, col)

                data.append(month)
                data.append(str(year))
                data.append(str(order))
                notes = worksheet[f'V{row}'].value

                if notes:
                    data.append(notes)

                total_data_.append(data)

        else:
            empty_rows += 1

        row += 1


file_name = find_xlsx_file()
if file_name is None:
    exit()

workbook = load_workbook(file_name, data_only=True)
workbook.create_sheet('Data')
total_data = [['Region', 'FBPN', 'Type', 'Month', 'Year', 'Planned_Orders', 'Notes']]

for sheet_name in ['NORAM', 'EMEA', 'LATAM', 'APAC']:
    process_data(workbook, sheet_name, total_data)

write_data(workbook, total_data)

save_file(workbook, file_name)
