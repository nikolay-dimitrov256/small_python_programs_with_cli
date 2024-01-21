import json
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

from sudoku import Sudoku


def load_data():
    if os.path.exists('sudoku log.txt'):
        with open('sudoku log.txt', 'r') as file:
            data = json.loads(file.read())
            last_sudoku = int(data['last sudoku'])
            last_sheet = int(data['last sheet'])
    else:
        last_sudoku = 0
        last_sheet = 0

    return last_sudoku, last_sheet


def save_log(last_sudoku: int, last_sheet: int):
    data = {'last sudoku': last_sudoku, 'last sheet': last_sheet}
    with open('sudoku log.txt', 'w') as file:
        file.write(json.dumps(data))


def write_sudoku_on_sheet(start_row: int):
    global sudoku_number
    sudoku_number += 1

    thin_border = Side(border_style='thin', color='000000')
    thick_border = Side(border_style='thick', color='000000')
    regular_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    bottom_border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thick_border)
    right_border = Border(left=thin_border, right=thick_border, top=thin_border, bottom=thin_border)
    bottom_right_border = Border(left=thin_border, right=thick_border, top=thin_border, bottom=thick_border)
    top_border = Border(left=thin_border, right=thin_border, top=thick_border, bottom=thin_border)
    left_border = Border(left=thick_border, right=thin_border, top=thin_border, bottom=thin_border)
    top_left_border = Border(left=thick_border, right=thin_border, top=thick_border, bottom=thin_border)
    top_right_border = Border(left=thin_border, right=thick_border, top=thick_border, bottom=thin_border)
    bottom_left_border = Border(left=thick_border, right=thin_border, top=thin_border, bottom=thick_border)

    sudoku = Sudoku(str(sudoku_number))

    sheet.merge_cells(f'A{start_row}:I{start_row}')
    sheet[f'A{start_row}'] = sudoku.name
    sheet[f'A{start_row}'].alignment = Alignment(horizontal='center', vertical='center')
    sheet[f'A{start_row}'].font = Font(size=20, bold=True)

    for row in range(9):
        for col in range(9):
            value = sudoku.player_board[row][col]
            col_letter = get_column_letter(col + 1)
            sheet_row = start_row + row + 1
            if value != 0:
                sheet[f'{col_letter}{sheet_row}'] = value
                sheet[f'{col_letter}{sheet_row}'].font = Font(size=20, bold=True)
            else:
                sheet[f'{col_letter}{sheet_row}'].font = Font(size=20, bold=True, color='003AFF')

            sheet[f'{col_letter}{sheet_row}'].alignment = Alignment(horizontal='center', vertical='center')

            if row % 3 == 0 and col % 3 == 0:
                sheet[f'{col_letter}{sheet_row}'].border = top_left_border
            elif row % 3 == 0 and col % 3 == 1:
                sheet[f'{col_letter}{sheet_row}'].border = top_border
            elif row % 3 == 0 and col % 3 == 2:
                sheet[f'{col_letter}{sheet_row}'].border = top_right_border
            elif row % 3 == 1 and col % 3 == 0:
                sheet[f'{col_letter}{sheet_row}'].border = left_border
            elif row % 3 == 1 and col % 3 == 1:
                sheet[f'{col_letter}{sheet_row}'].border = regular_border
            elif row % 3 == 1 and col % 3 == 2:
                sheet[f'{col_letter}{sheet_row}'].border = right_border
            elif row % 3 == 2 and col % 3 == 0:
                sheet[f'{col_letter}{sheet_row}'].border = bottom_left_border
            elif row % 3 == 2 and col % 3 == 1:
                sheet[f'{col_letter}{sheet_row}'].border = bottom_border
            elif row % 3 == 2 and col % 3 == 2:
                sheet[f'{col_letter}{sheet_row}'].border = bottom_right_border


sudoku_number, sheet_number = load_data()

sheet_number += 1

wb = Workbook()
sheet = wb.active

for column in 'ABCDEFGHI':
    sheet.column_dimensions[column].width = 5

for row in range(1, 22):
    sheet.row_dimensions[row].height = 25

for i in range(1, 13, 11):
    write_sudoku_on_sheet(i)

wb.save(f'sudoku {sheet_number}.xlsx')
save_log(sudoku_number, sheet_number)
